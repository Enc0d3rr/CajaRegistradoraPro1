from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QMessageBox, QComboBox, QGroupBox
)
from PyQt6.QtGui import QPalette, QColor
from PyQt6.QtCore import Qt
import os
import csv
from datetime import datetime
from utils.helpers import formato_moneda_mx  

class ExportDialog(QDialog):
    def __init__(self, db_manager, report_type, date_range, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.report_type = report_type
        self.date_range = date_range
        self.setWindowTitle(f"Exportar Reporte de {report_type.capitalize()}")
        self.setGeometry(300, 200, 450, 250)
        
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#ecf0f1"))
        self.setPalette(palette)
        
        layout = QVBoxLayout()
        
        # Información del reporte
        info_group = QGroupBox("Información del Reporte")
        info_layout = QVBoxLayout()
        info_layout.addWidget(QLabel(f"Tipo: Reporte de {report_type}"))
        info_layout.addWidget(QLabel(f"Desde: {date_range['desde']}"))
        info_layout.addWidget(QLabel(f"Hasta: {date_range['hasta']}"))
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Selección de formato
        format_group = QGroupBox("Formato de Exportación")
        format_layout = QVBoxLayout()
        format_layout.addWidget(QLabel("Seleccione el formato:"))
        self.format_combo = QComboBox()
        self.format_combo.addItems([
            "PDF - Documento imprimible", 
            "Excel - Hoja de cálculo", 
            "CSV - Datos para análisis"
        ])
        format_layout.addWidget(self.format_combo)
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_export = QPushButton("💾 Exportar")
        btn_export.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 10px;")
        btn_export.clicked.connect(self.exportar_reporte)
        buttons_layout.addWidget(btn_export)
        
        btn_cancel = QPushButton("❌ Cancelar")
        btn_cancel.setStyleSheet("background-color: #e74c3c; color: white; font-weight: bold; padding: 10px;")
        btn_cancel.clicked.connect(self.reject)
        buttons_layout.addWidget(btn_cancel)
        
        layout.addLayout(buttons_layout)
        self.setLayout(layout)
    
    def verificar_datos_reporte(self):
        """Verificar si hay datos para el reporte"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                fecha_desde = self.date_range['desde'].split(' ')[0] + ' 00:00:00'
                fecha_hasta = self.date_range['hasta'].split(' ')[0] + ' 23:59:59'
                
                if self.report_type == "ventas":
                    cursor.execute("SELECT COUNT(*) FROM ventas WHERE fecha BETWEEN ? AND ?", (fecha_desde, fecha_hasta))
                else:
                    cursor.execute("SELECT COUNT(*) FROM cierres_caja WHERE fecha_apertura BETWEEN ? AND ?", (fecha_desde, fecha_hasta))
                
                count = cursor.fetchone()[0]
                return count > 0
                
        except Exception as e:
            return False
    
    def normalizar_fechas_consulta(self):
        """Normalizar formato de fechas para consultas SQL - CON DIAGNÓSTICO"""
        try:
            print(f"🔍 Normalizando fechas. Original: {self.date_range}")
            
            # Verificar que date_range tenga la estructura correcta
            if not isinstance(self.date_range, dict) or 'desde' not in self.date_range or 'hasta' not in self.date_range:
                print("❌ Estructura de date_range incorrecta, usando fechas por defecto")
                return {
                    'desde': '2000-01-01 00:00:00',
                    'hasta': '2030-12-31 23:59:59'
                }
            
            desde_str = str(self.date_range['desde'])
            hasta_str = str(self.date_range['hasta'])
            
            print(f"📅 Fechas originales - Desde: '{desde_str}', Hasta: '{hasta_str}'")
            
            # Si son solo fechas (sin hora), agregar hora
            if ' ' not in desde_str:
                fecha_desde = f"{desde_str} 00:00:00"
            else:
                fecha_desde = desde_str
                
            if ' ' not in hasta_str:
                fecha_hasta = f"{hasta_str} 23:59:59"
            else:
                fecha_hasta = hasta_str
            
            print(f"📅 Fechas normalizadas - Desde: '{fecha_desde}', Hasta: '{fecha_hasta}'")
            
            return {
                'desde': fecha_desde,
                'hasta': fecha_hasta
            }
                
        except Exception as e:
            print(f"❌ Error normalizando fechas: {e}")
            # Fechas por defecto muy amplias para asegurar que haya datos
            return {
                'desde': '2000-01-01 00:00:00',
                'hasta': '2030-12-31 23:59:59'
            }
        
    def diagnosticar_exportacion(self):
        """Función de diagnóstico para ver qué está pasando"""
        print("=== DIAGNÓSTICO DE EXPORTACIÓN ===")
        
        # 1. Verificar conexión a la base de datos
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar tablas existentes
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tablas = cursor.fetchall()
                print("📊 Tablas en la base de datos:", [t[0] for t in tablas])
                
                # Verificar fechas que se están usando
                fechas = self.normalizar_fechas_consulta()
                print("📅 Fechas normalizadas:", fechas)
                
                # Verificar datos en ventas
                if self.report_type == "ventas":
                    cursor.execute("SELECT COUNT(*) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                                (fechas['desde'], fechas['hasta']))
                    count_ventas = cursor.fetchone()[0]
                    print(f"🛒 Ventas en el período: {count_ventas}")
                    
                    # Ver algunas ventas de ejemplo
                    cursor.execute("SELECT id, fecha, total FROM ventas ORDER BY fecha DESC LIMIT 3")
                    ventas_ejemplo = cursor.fetchall()
                    print("📝 Últimas 3 ventas:", ventas_ejemplo)
                    
                else:  # cierres
                    cursor.execute("SELECT COUNT(*) FROM cierres_caja WHERE fecha_apertura BETWEEN ? AND ?", 
                                (fechas['desde'], fechas['hasta']))
                    count_cierres = cursor.fetchone()[0]
                    print(f"💰 Cierres en el período: {count_cierres}")
                    
                    # Ver algunos cierres de ejemplo
                    cursor.execute("SELECT id, fecha_apertura, total_ventas FROM cierres_caja ORDER BY fecha_apertura DESC LIMIT 3")
                    cierres_ejemplo = cursor.fetchall()
                    print("📝 Últimos 3 cierres:", cierres_ejemplo)
                    
        except Exception as e:
            print(f"❌ Error en diagnóstico: {e}")
    
    def exportar_reporte(self):
        formato = self.format_combo.currentText()
        reports_dir = "Reportes"
        os.makedirs(reports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"{reports_dir}/{self.report_type}_{timestamp}"
        
        try:
            # ✅ EJECUTAR DIAGNÓSTICO PRIMERO
            self.diagnosticar_exportacion()
            
            if "PDF" in formato:
                filename = f"{base_filename}.pdf"
                self.exportar_pdf(filename)
                
            elif "Excel" in formato:
                filename = f"{base_filename}.xlsx"
                self.exportar_excel(filename)
                
            elif "CSV" in formato:
                filename = f"{base_filename}.csv"
                self.exportar_csv(filename)
            
            QMessageBox.information(self, "✅ Éxito", 
                f"Reporte exportado correctamente\n\nArchivo: {os.path.basename(filename)}")
            
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo exportar: {str(e)}")
    
    def exportar_pdf(self, filename):
        """Exportar a PDF - VERSIÓN MEJORADA QUE USA REPORTLAB REAL"""
        try:
            print(f"🔄 Exportando PDF a: {filename}")
            
            # Verificar explícitamente si reportlab está disponible
            try:
                import reportlab
                print("✅ ReportLab está instalado, generando PDF real...")
                
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                
                # Crear documento PDF real
                doc = SimpleDocTemplate(filename, pagesize=A4)
                elements = []
                styles = getSampleStyleSheet()
                
                # Título
                title_style = getSampleStyleSheet()['Heading1']
                elements.append(Paragraph(f"REPORTE DE {self.report_type.upper()}", title_style))
                elements.append(Spacer(1, 20))
                
                # Información del reporte
                normal_style = styles['Normal']
                elements.append(Paragraph(f"<b>Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
                elements.append(Paragraph(f"<b>Período:</b> {self.date_range['desde']} a {self.date_range['hasta']}", normal_style))
                elements.append(Spacer(1, 20))
                
                # Contenido específico del reporte
                if self.report_type == "ventas":
                    content = self._generar_contenido_ventas_pdf()
                else:
                    content = self._generar_contenido_cierres_pdf()
                
                elements.extend(content)
                
                # Construir PDF
                doc.build(elements)
                print(f"✅ PDF real generado correctamente: {filename}")
                
                QMessageBox.information(self, "✅ PDF Exportado", 
                    f"PDF con formato profesional generado correctamente\n\nArchivo: {os.path.basename(filename)}")
                
                return filename
                
            except ImportError as e:
                print(f"⚠️ ReportLab no disponible: {e}")
                raise ImportError("ReportLab no está instalado")
                
        except ImportError:
            # Fallback a texto plano solo si ReportLab no está instalado
            print("🔄 Usando fallback a texto plano para PDF")
            return self._exportar_pdf_fallback(filename)
        except Exception as e:
            print(f"❌ Error generando PDF real: {e}")
            # Fallback a texto plano en caso de error
            return self._exportar_pdf_fallback(filename)
        
    def _exportar_pdf_fallback(self, filename):
        """Fallback a texto plano para PDF"""
        txt_filename = filename.replace('.pdf', '.txt')
        
        with open(txt_filename, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("SISTEMA DE CAJA REGISTRADORA\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"REPORTE DE {self.report_type.upper()}\n")
            f.write(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Período: {self.date_range['desde']} a {self.date_range['hasta']}\n\n")
            
            if self.report_type == "ventas":
                self._exportar_ventas_texto(f)
            else:
                self._exportar_cierres_texto(f)
        
        print(f"✅ PDF fallback generado como texto: {txt_filename}")
        return txt_filename
    
    def _generar_contenido_ventas_pdf(self):
        """Generar contenido para PDF de ventas - VERSIÓN SIMPLIFICADA Y FUNCIONAL"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        elements = []
        styles = getSampleStyleSheet()
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                fechas = self.normalizar_fechas_consulta()
                
                # Resumen
                cursor.execute("SELECT COUNT(*), SUM(total) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                            (fechas['desde'], fechas['hasta']))
                resultado = cursor.fetchone()
                total_ventas = resultado[0] if resultado else 0
                monto_total = resultado[1] if resultado else 0
                
                elements.append(Paragraph(f"<b>Total de ventas:</b> {total_ventas}", styles['Normal']))
                elements.append(Paragraph(f"<b>Monto total:</b> {formato_moneda_mx(monto_total)}", styles['Normal']))
                elements.append(Spacer(1, 15))
                
                if total_ventas > 0:
                    # Detalle de ventas (limitado para no hacer el PDF muy grande)
                    cursor.execute("""
                        SELECT v.id, v.fecha, v.total, v.metodo_pago, u.nombre
                        FROM ventas v 
                        JOIN usuarios u ON v.usuario_id = u.id 
                        WHERE v.fecha BETWEEN ? AND ?
                        ORDER BY v.fecha DESC
                        LIMIT 30
                    """, (fechas['desde'], fechas['hasta']))
                    
                    ventas = cursor.fetchall()
                    
                    # Crear tabla simple
                    data = [['ID', 'Fecha', 'Total', 'Método', 'Vendedor']]
                    for venta in ventas:
                        id_venta, fecha, total, metodo, vendedor = venta
                        # Formatear fecha para que sea más corta
                        fecha_corta = fecha.split()[0] if ' ' in fecha else fecha
                        data.append([
                            str(id_venta),
                            fecha_corta,
                            formato_moneda_mx(total),
                            metodo,
                            vendedor
                        ])
                    
                    if len(data) > 1:  # Si hay datos además del encabezado
                        table = Table(data, colWidths=[50, 80, 70, 80, 100])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F3F3')),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                        ]))
                        elements.append(table)
                    else:
                        elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
                
        except Exception as e:
            print(f"❌ Error generando contenido PDF: {e}")
            elements.append(Paragraph(f"Error al generar contenido: {str(e)}", styles['Normal']))
        
        return elements
    
    def _generar_contenido_cierres_pdf(self):
        """Generar contenido para PDF de cierres"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        elements = []
        styles = getSampleStyleSheet()
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                fechas = self.normalizar_fechas_consulta()
                
                # Resumen
                cursor.execute("SELECT COUNT(*), SUM(total_ventas), AVG(diferencia) FROM cierres_caja WHERE fecha_apertura BETWEEN ? AND ?", 
                             (fechas['desde'], fechas['hasta']))
                resultado = cursor.fetchone()
                total_cierres = resultado[0] if resultado else 0
                ventas_total = resultado[1] if resultado else 0
                diff_promedio = resultado[2] if resultado else 0
                
                elements.append(Paragraph(f"<b>Total de cierres:</b> {total_cierres}", styles['Normal']))
                elements.append(Paragraph(f"<b>Ventas totales:</b> {formato_moneda_mx(ventas_total)}", styles['Normal']))
                elements.append(Paragraph(f"<b>Diferencia promedio:</b> {formato_moneda_mx(diff_promedio)}", styles['Normal']))
                elements.append(Spacer(1, 15))
                
                if total_cierres > 0:
                    # Detalle de cierres
                    cursor.execute("""
                        SELECT c.fecha_apertura, u.nombre, c.monto_inicial, c.ventas_efectivo,
                               c.ventas_tarjeta, c.total_ventas, c.diferencia
                        FROM cierres_caja c
                        JOIN usuarios u ON c.usuario_id = u.id
                        WHERE c.fecha_apertura BETWEEN ? AND ?
                        ORDER BY c.fecha_apertura DESC
                    """, (fechas['desde'], fechas['hasta']))
                    
                    cierres = cursor.fetchall()
                    
                    # Crear tabla
                    data = [['Fecha', 'Usuario', 'Monto Inicial', 'Ventas Efectivo', 'Ventas Tarjeta', 'Total', 'Diferencia']]
                    for cierre in cierres:
                        fecha, usuario, inicial, efectivo, tarjeta, total, diff = cierre
                        fecha_corta = fecha.split()[0] if ' ' in fecha else fecha
                        data.append([
                            fecha_corta,
                            usuario,
                            formato_moneda_mx(inicial),
                            formato_moneda_mx(efectivo),
                            formato_moneda_mx(tarjeta),
                            formato_moneda_mx(total),
                            formato_moneda_mx(diff)
                        ])
                    
                    if len(data) > 1:
                        table = Table(data, colWidths=[80, 80, 70, 70, 70, 70, 70])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F3F3')),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                        ]))
                        elements.append(table)
                    else:
                        elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
                
        except Exception as e:
            print(f"❌ Error generando contenido PDF cierres: {e}")
            elements.append(Paragraph(f"Error al generar contenido: {str(e)}", styles['Normal']))
        
        return elements
    
    def exportar_excel(self, filename):
        """Exportar a Excel - VERSIÓN CON OPENPYXL REAL"""
        try:
            print(f"🔄 Exportando Excel a: {filename}")
            
            # Verificar explícitamente si openpyxl está disponible
            try:
                import openpyxl
                print("✅ OpenPyXL está instalado, generando Excel real...")
                
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
                
                # Crear libro de Excel real
                wb = Workbook()
                ws = wb.active
                ws.title = f"Reporte {self.report_type}"
                
                # Estilos
                title_font = Font(bold=True, size=14)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                normal_font = Font(size=10)
                
                # Título
                ws['A1'] = f"REPORTE DE {self.report_type.upper()}"
                ws['A1'].font = title_font
                ws.merge_cells('A1:G1')
                
                # Información del reporte
                ws['A3'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A4'] = f"Período: {self.date_range['desde']} a {self.date_range['hasta']}"
                
                if self.report_type == "ventas":
                    self._generar_excel_ventas(ws, header_font, header_fill, normal_font)
                else:
                    self._generar_excel_cierres(ws, header_font, header_fill, normal_font)
                
                # Ajustar anchos de columna
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Guardar archivo Excel real
                wb.save(filename)
                print(f"✅ Excel real generado correctamente: {filename}")
                
                QMessageBox.information(self, "✅ Excel Exportado", 
                    f"Archivo Excel (.xlsx) generado correctamente\n\nArchivo: {os.path.basename(filename)}")
                
                return filename
                
            except ImportError:
                print("⚠️ OpenPyXL no disponible, usando CSV como fallback")
                raise ImportError("OpenPyXL no está instalado")
                
        except ImportError:
            # Fallback a CSV solo si OpenPyXL no está instalado
            print("🔄 Usando fallback a CSV para Excel")
            return self._exportar_excel_fallback(filename)
        except Exception as e:
            print(f"❌ Error generando Excel real: {e}")
            # Fallback a CSV en caso de error
            return self._exportar_excel_fallback(filename)
        
    def _generar_excel_ventas(self, ws, header_font, header_fill, normal_font):
        """Generar contenido de ventas para Excel"""
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            fechas = self.normalizar_fechas_consulta()
            
            # Resumen
            cursor.execute("SELECT COUNT(*), SUM(total) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                         (fechas['desde'], fechas['hasta']))
            resultado = cursor.fetchone()
            total_ventas = resultado[0] if resultado else 0
            monto_total = resultado[1] if resultado else 0
            
            ws['A6'] = "RESUMEN GENERAL"
            ws['A6'].font = header_font
            ws['A7'] = f"Total de ventas: {total_ventas}"
            ws['A8'] = f"Monto total: {formato_moneda_mx(monto_total)}"
            
            if total_ventas > 0:
                # Encabezados de tabla
                headers = ['ID', 'Fecha', 'Total', 'IVA', 'Método Pago', 'Vendedor']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=10, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Datos de ventas
                cursor.execute("""
                    SELECT v.id, v.fecha, v.total, v.iva, v.metodo_pago, u.nombre
                    FROM ventas v 
                    JOIN usuarios u ON v.usuario_id = u.id 
                    WHERE v.fecha BETWEEN ? AND ?
                    ORDER BY v.fecha DESC
                """, (fechas['desde'], fechas['hasta']))
                
                for row, venta in enumerate(cursor.fetchall(), 11):
                    for col, value in enumerate(venta, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = normal_font

    def _generar_excel_cierres(self, ws, header_font, header_fill, normal_font):
        """Generar contenido de cierres para Excel"""
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            fechas = self.normalizar_fechas_consulta()
            
            # Resumen
            cursor.execute("SELECT COUNT(*), SUM(total_ventas), AVG(diferencia) FROM cierres_caja WHERE fecha_apertura BETWEEN ? AND ?", 
                         (fechas['desde'], fechas['hasta']))
            resultado = cursor.fetchone()
            total_cierres = resultado[0] if resultado else 0
            ventas_total = resultado[1] if resultado else 0
            diff_promedio = resultado[2] if resultado else 0
            
            ws['A6'] = "RESUMEN GENERAL"
            ws['A6'].font = header_font
            ws['A7'] = f"Total de cierres: {total_cierres}"
            ws['A8'] = f"Ventas totales: {formato_moneda_mx(ventas_total)}"
            ws['A9'] = f"Diferencia promedio: {formato_moneda_mx(diff_promedio)}"
            
            if total_cierres > 0:
                # Encabezados de tabla
                headers = ['Fecha', 'Usuario', 'Monto Inicial', 'Ventas Efectivo', 'Ventas Tarjeta', 'Total Ventas', 'Diferencia']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=11, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Datos de cierres
                cursor.execute("""
                    SELECT c.fecha_apertura, u.nombre, c.monto_inicial, c.ventas_efectivo,
                           c.ventas_tarjeta, c.total_ventas, c.diferencia
                    FROM cierres_caja c
                    JOIN usuarios u ON c.usuario_id = u.id
                    WHERE c.fecha_apertura BETWEEN ? AND ?
                    ORDER BY c.fecha_apertura DESC
                """, (fechas['desde'], fechas['hasta']))
                
                for row, cierre in enumerate(cursor.fetchall(), 12):
                    for col, value in enumerate(cierre, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = normal_font

    def _exportar_excel_fallback(self, filename):
        """Fallback a CSV para Excel"""
        csv_filename = filename.replace('.xlsx', '.csv')
        
        with open(csv_filename, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            
            writer.writerow(['Reporte exportado desde Sistema de Caja Registradora'])
            writer.writerow([f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([f'Tipo: {self.report_type}'])
            writer.writerow([f'Período: {self.date_range["desde"]} a {self.date_range["hasta"]}'])
            writer.writerow([])
            
            if self.report_type == "ventas":
                self._exportar_ventas_csv(writer)
            else:
                self._exportar_cierres_csv(writer)
        
        print(f"✅ Excel fallback generado como CSV: {csv_filename}")
        
        QMessageBox.information(self, "📊 Excel (Fallback)", 
            "Se generó un archivo CSV. Para formato .xlsx real, instale: pip install openpyxl")
        
        return csv_filename
    
    def _exportar_ventas_texto(self, f):
        """Datos de ventas para texto plano - CORREGIDO"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                fechas = self.normalizar_fechas_consulta()
                
                # ✅ CORRECCIÓN: Consulta sin la columna 'cliente' que no existe
                cursor.execute("SELECT COUNT(*), SUM(total) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                            (fechas['desde'], fechas['hasta']))
                
                resultado = cursor.fetchone()
                total_ventas = resultado[0] if resultado else 0
                monto_total = resultado[1] if resultado else 0
                
                f.write("RESUMEN GENERAL\n")
                f.write("-" * 40 + "\n")
                f.write(f"Total de ventas: {total_ventas}\n")
                f.write(f"Monto total: {formato_moneda_mx(monto_total)}\n\n")
                
                if total_ventas == 0:
                    f.write("NO SE ENCONTRARON VENTAS EN EL PERÍODO SELECCIONADO\n")
                    return
                
                # ✅ CORRECCIÓN: Consulta sin la columna 'cliente'
                f.write("DETALLE DE VENTAS\n")
                f.write("-" * 40 + "\n")
                
                cursor.execute("""
                    SELECT v.id, v.fecha, v.total, v.iva, v.metodo_pago, u.nombre
                    FROM ventas v 
                    JOIN usuarios u ON v.usuario_id = u.id 
                    WHERE v.fecha BETWEEN ? AND ?
                    ORDER BY v.fecha DESC
                """, (fechas['desde'], fechas['hasta']))
                
                for id_venta, fecha, total, iva, metodo, vendedor in cursor.fetchall():
                    f.write(f"Venta #{id_venta} - {fecha}\n")
                    f.write(f"  Total: {formato_moneda_mx(total)}\n")
                    f.write(f"  IVA: {formato_moneda_mx(iva)}\n")
                    f.write(f"  Método: {metodo}\n")
                    f.write(f"  Vendedor: {vendedor}\n")
                    f.write("-" * 30 + "\n")
                    
        except Exception as e:
            f.write(f"ERROR AL EXPORTAR VENTAS: {str(e)}\n")
    
    def _exportar_cierres_texto(self, f):
        """Datos de cierres para texto plano - CORREGIDO"""
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            fechas = self.normalizar_fechas_consulta()
        
            # ✅ CORRECCIÓN: Usar fetchone() en lugar de fetchall() para totales
            cursor.execute("SELECT COUNT(*), SUM(total_ventas), AVG(diferencia) FROM cierres_caja WHERE fecha_apertura BETWEEN ? AND ?", 
                        (fechas['desde'], fechas['hasta']))
            
            resultado = cursor.fetchone()  # ✅ Cambiado a fetchone()
            
            total_cierres = resultado[0] if resultado else 0
            ventas_total = resultado[1] if resultado else 0
            diff_promedio = resultado[2] if resultado else 0
        
            # ✅ CORRECCIÓN: Usar formato_moneda_mx
            ventas_total_formateado = formato_moneda_mx(ventas_total)
            diff_promedio_formateado = formato_moneda_mx(diff_promedio)
            
            f.write(f"Cierres realizados: {total_cierres}\n")
            f.write(f"Ventas totales: {ventas_total_formateado}\n")
            f.write(f"Diferencia promedio: {diff_promedio_formateado}\n\n")
        
            f.write("DETALLES DE CIERRES:\n")
            f.write("=" * 50 + "\n")
        
            cursor.execute("""
                SELECT c.fecha_apertura, u.nombre, c.monto_inicial, c.ventas_efectivo,
                    c.ventas_tarjeta, c.total_ventas, c.diferencia
                FROM cierres_caja c
                JOIN usuarios u ON c.usuario_id = u.id
                WHERE c.fecha_apertura BETWEEN ? AND ?
                ORDER BY c.fecha_apertura DESC
            """, (fechas['desde'], fechas['hasta']))
        
            for fecha, usuario, inicial, efectivo, tarjeta, total, diff in cursor.fetchall():
                # ✅ CORRECCIÓN: Usar formato_moneda_mx
                inicial_formateado = formato_moneda_mx(inicial or 0)
                efectivo_formateado = formato_moneda_mx(efectivo or 0)
                tarjeta_formateado = formato_moneda_mx(tarjeta or 0)
                total_formateado = formato_moneda_mx(total or 0)
                diff_formateado = formato_moneda_mx(diff or 0)
                
                f.write(f"\nFecha: {fecha}\n")
                f.write(f"Usuario: {usuario}\n")
                f.write(f"Efectivo inicial: {inicial_formateado}\n")
                f.write(f"Ventas efectivo: {efectivo_formateado}\n")
                f.write(f"Ventas tarjeta: {tarjeta_formateado}\n")
                f.write(f"Total ventas: {total_formateado}\n")
                f.write(f"Diferencia: {diff_formateado}\n")
                f.write("-" * 30 + "\n")

    def _exportar_ventas_excel(self, f):
        """Datos de ventas para Excel - CORREGIDO"""
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            fechas = self.normalizar_fechas_consulta()
            
            cursor.execute("SELECT COUNT(*), SUM(total) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                        (fechas['desde'], fechas['hasta']))
            total_ventas, monto_total = cursor.fetchone()
            
            # ✅ CORRECCIÓN: Usar formato_moneda_mx
            monto_total_formateado = formato_moneda_mx(monto_total or 0)
            f.write(f"Total Ventas,{total_ventas or 0}\n")
            f.write(f"Monto Total,{monto_total_formateado}\n\n")
            
            # ✅ CORRECCIÓN: Sin columna 'cliente'
            f.write("Detalle de Ventas\n")
            f.write("ID,Fecha,Total,IVA,Método Pago,Usuario\n")
            
            cursor.execute("""
                SELECT v.id, v.fecha, v.total, v.iva, v.metodo_pago, u.nombre
                FROM ventas v 
                JOIN usuarios u ON v.usuario_id = u.id 
                WHERE v.fecha BETWEEN ? AND ?
                ORDER BY v.fecha
            """, (fechas['desde'], fechas['hasta']))
            
            for id_venta, fecha, total, iva, metodo, usuario in cursor.fetchall():
                # ✅ CORRECCIÓN: Usar formato_moneda_mx
                total_formateado = formato_moneda_mx(total)
                f.write(f"{id_venta},{fecha},{total_formateado},{iva},{metodo},{usuario}\n")
    
    def _exportar_cierres_excel(self, f):
        """Datos de cierres para Excel"""
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            fechas = self.normalizar_fechas_consulta()
            
            f.write("Detalle de Cierres de Caja\n")
            f.write("Fecha Apertura,Usuario,Monto Inicial,Ventas Efectivo,Ventas Tarjeta,Total Ventas,Diferencia\n")
            
            cursor.execute("""
                SELECT c.fecha_apertura, u.nombre, c.monto_inicial, c.ventas_efectivo,
                       c.ventas_tarjeta, c.total_ventas, c.diferencia
                FROM cierres_caja c
                JOIN usuarios u ON c.usuario_id = u.id
                WHERE c.fecha_apertura BETWEEN ? AND ?
                ORDER BY c.fecha_apertura
            """, (fechas['desde'], fechas['hasta']))
            
            for fecha, usuario, inicial, efectivo, tarjeta, total, diff in cursor.fetchall():
                # ✅ CORRECCIÓN: Usar formato_moneda_mx
                inicial_formateado = formato_moneda_mx(inicial)
                efectivo_formateado = formato_moneda_mx(efectivo)
                tarjeta_formateado = formato_moneda_mx(tarjeta)
                total_formateado = formato_moneda_mx(total)
                diff_formateado = formato_moneda_mx(diff)
                f.write(f"{fecha},{usuario},{inicial_formateado},{efectivo_formateado},{tarjeta_formateado},{total_formateado},{diff_formateado}\n")
    
    def exportar_csv(self, filename):
        """Exportar a CSV - VERSIÓN COMPLETAMENTE CORREGIDA"""
        try:
            print(f"🔄 Exportando CSV a: {filename}")
            
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                
                # Encabezado del reporte
                writer.writerow(['=== SISTEMA DE CAJA REGISTRADORA ==='])
                writer.writerow([f'Reporte de {self.report_type.title()}'])
                writer.writerow([f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                writer.writerow([f'Período: {self.date_range["desde"]} a {self.date_range["hasta"]}'])
                writer.writerow([])  # Línea vacía
                
                if self.report_type == "ventas":
                    self._exportar_ventas_csv(writer)
                else:
                    self._exportar_cierres_csv(writer)
                    
            print(f"✅ CSV exportado exitosamente: {filename}")
            return filename
                        
        except Exception as e:
            print(f"❌ Error exportando CSV: {e}")
            raise Exception(f"Error CSV: {str(e)}")
    
    def _exportar_ventas_csv(self, writer):
        """Datos de ventas para CSV - CORREGIDO"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                fechas = self.normalizar_fechas_consulta()
                
                print(f"🔍 Buscando ventas entre: {fechas['desde']} y {fechas['hasta']}")
                
                # 1. Totales generales
                cursor.execute("SELECT COUNT(*), SUM(total) FROM ventas WHERE fecha BETWEEN ? AND ?", 
                            (fechas['desde'], fechas['hasta']))
                
                resultado = cursor.fetchone()
                total_ventas = resultado[0] if resultado else 0
                monto_total = resultado[1] if resultado else 0
                
                print(f"📊 Encontradas {total_ventas} ventas, total: {monto_total}")
                
                writer.writerow(['RESUMEN GENERAL'])
                writer.writerow(['Total de ventas:', total_ventas])
                writer.writerow(['Monto total:', formato_moneda_mx(monto_total)])
                writer.writerow([])
                
                if total_ventas == 0:
                    writer.writerow(['NO SE ENCONTRARON VENTAS EN EL PERÍODO SELECCIONADO'])
                    return
                
                # 2. Por método de pago
                writer.writerow(['DISTRIBUCIÓN POR MÉTODO DE PAGO'])
                writer.writerow(['Método', 'Cantidad', 'Monto Total'])
                
                cursor.execute("""
                    SELECT metodo_pago, COUNT(*), SUM(total) 
                    FROM ventas 
                    WHERE fecha BETWEEN ? AND ?
                    GROUP BY metodo_pago
                """, (fechas['desde'], fechas['hasta']))
                
                for metodo, cantidad, monto in cursor.fetchall():
                    writer.writerow([metodo, cantidad, formato_moneda_mx(monto or 0)])
                
                writer.writerow([])
                
                # 3. ✅ CORRECCIÓN: Detalle completo de ventas sin columna 'cliente'
                writer.writerow(['DETALLE COMPLETO DE VENTAS'])
                writer.writerow(['ID', 'Fecha', 'Total', 'IVA', 'Método Pago', 'Vendedor'])
                
                cursor.execute("""
                    SELECT v.id, v.fecha, v.total, v.iva, v.metodo_pago, u.nombre
                    FROM ventas v 
                    JOIN usuarios u ON v.usuario_id = u.id 
                    WHERE v.fecha BETWEEN ? AND ?
                    ORDER BY v.fecha DESC
                """, (fechas['desde'], fechas['hasta']))
                
                ventas = cursor.fetchall()
                print(f"📝 Exportando {len(ventas)} ventas detalladas")
                
                for venta in ventas:
                    id_venta, fecha, total, iva, metodo, vendedor = venta
                    writer.writerow([
                        id_venta, 
                        fecha, 
                        formato_moneda_mx(total), 
                        formato_moneda_mx(iva), 
                        metodo, 
                        vendedor
                    ])
                
                print("✅ Datos de ventas escritos correctamente en CSV")
                
        except Exception as e:
            print(f"❌ Error en exportación CSV ventas: {e}")
            writer.writerow(['ERROR AL EXPORTAR VENTAS:', str(e)])
    
    def _exportar_cierres_csv(self, writer):
        """Datos de cierres para CSV - VERSIÓN CORREGIDA"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                fechas = self.normalizar_fechas_consulta()
                
                print(f"🔍 Buscando cierres entre: {fechas['desde']} y {fechas['hasta']}")
                
                writer.writerow(['REPORTE DE CIERRES DE CAJA'])
                writer.writerow([])
                
                # 1. Resumen general
                cursor.execute("""
                    SELECT COUNT(*), SUM(total_ventas), AVG(diferencia)
                    FROM cierres_caja 
                    WHERE fecha_apertura BETWEEN ? AND ?
                """, (fechas['desde'], fechas['hasta']))
                
                resultado = cursor.fetchone()
                total_cierres = resultado[0] if resultado else 0
                ventas_totales = resultado[1] if resultado else 0
                diff_promedio = resultado[2] if resultado else 0
                
                print(f"📊 Encontrados {total_cierres} cierres")
                
                writer.writerow(['RESUMEN GENERAL'])
                writer.writerow(['Total de cierres:', total_cierres])
                writer.writerow(['Ventas totales:', formato_moneda_mx(ventas_totales)])
                writer.writerow(['Diferencia promedio:', formato_moneda_mx(diff_promedio)])
                writer.writerow([])
                
                if total_cierres == 0:
                    writer.writerow(['NO SE ENCONTRARON CIERRES EN EL PERÍODO SELECCIONADO'])
                    return
                
                # 2. Detalle de cierres
                writer.writerow(['DETALLE DE CIERRES'])
                writer.writerow(['Fecha', 'Usuario', 'Monto Inicial', 'Ventas Efectivo', 
                            'Ventas Tarjeta', 'Ventas Transferencia', 'Total Ventas', 'Diferencia'])
                
                cursor.execute("""
                    SELECT c.fecha_apertura, u.nombre, c.monto_inicial, c.ventas_efectivo,
                        c.ventas_tarjeta, c.ventas_transferencia, c.total_ventas, c.diferencia
                    FROM cierres_caja c
                    JOIN usuarios u ON c.usuario_id = u.id
                    WHERE c.fecha_apertura BETWEEN ? AND ?
                    ORDER BY c.fecha_apertura DESC
                """, (fechas['desde'], fechas['hasta']))
                
                cierres = cursor.fetchall()
                print(f"📝 Exportando {len(cierres)} cierres detallados")
                
                for cierre in cierres:
                    fecha, usuario, inicial, efectivo, tarjeta, transferencia, total, diff = cierre
                    writer.writerow([
                        fecha,
                        usuario,
                        formato_moneda_mx(inicial),
                        formato_moneda_mx(efectivo),
                        formato_moneda_mx(tarjeta),
                        formato_moneda_mx(transferencia),
                        formato_moneda_mx(total),
                        formato_moneda_mx(diff)
                    ])
                
                print("✅ Datos de cierres escritos correctamente en CSV")
                
        except Exception as e:
            print(f"❌ Error en exportación CSV cierres: {e}")
            writer.writerow(['ERROR AL EXPORTAR CIERRES:', str(e)])

    